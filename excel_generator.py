"""Excel file generator for debt analysis results."""

from datetime import datetime
from pathlib import Path
from typing import List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

from config import config
from csv_parser import DebtAccount
from debt_analyzer import CategorizationResult


# Category colors for highlighting (matching CSS)
CATEGORY_COLORS = {
    'MED': 'E0E0E0',        # Grey
    'MED READY': 'BDBDBD',  # Darker grey
    'PAY': 'FFE0B2',        # Orange
    'PAID': 'C5CAE9',       # Indigo
    'INS': 'C8E6C9',        # Light green
    'SMJ': 'E1BEE7',        # Purple
    'BAD': 'FFCDD2',        # Red
    'INV': 'B3E5FC',        # Light blue
    'STAFF': 'FFF9C4',      # Light yellow
}


def create_debt_excel(
    accounts: List[DebtAccount],
    results: List[CategorizationResult],
    output_path: Path = None
) -> Path:
    """
    Create an Excel file with debt analysis results.

    Args:
        accounts: List of DebtAccount objects from CSV
        results: List of CategorizationResult objects from analysis
        output_path: Optional output path. Defaults to output folder with timestamp.

    Returns:
        Path to the created Excel file
    """
    if output_path is None:
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        output_path = config.OUTPUT_FOLDER / f'debt_analysis_{timestamp}.xlsx'

    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Create results lookup by client_id
    results_map = {r.client_id: r for r in results}

    # Filter out accounts excluded from Excel (BAD, INV, etc.)
    filtered_accounts = []
    for account in accounts:
        result = results_map.get(account.client_id)
        if result and result.exclude_from_excel:
            continue  # Skip this account
        filtered_accounts.append(account)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Debt Analysis"

    # Define headers
    headers = [
        'ID',
        'Name',
        'Last Payment',
        'Current',
        '30 Days',
        '60 Days',
        '90 Days',
        'Total',
        'SMS',
        'EMAIL',
        'TYPE',
        'ADMIN NOTES',       # New column for practice manager notes
        'AI ANALYSIS NOTES'  # Renamed from NOTES
    ]

    # Column widths
    column_widths = [15, 35, 22, 12, 12, 12, 12, 12, 8, 8, 12, 25, 50]

    # Style definitions
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4A90D9', end_color='4A90D9', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Set column widths
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Write data rows (using filtered accounts)
    row_num = 2
    for account in filtered_accounts:
        result = results_map.get(account.client_id)

        # Format last payment
        if account.last_payment_date:
            last_payment = f"£{account.last_payment_amount:.2f} on {account.last_payment_date.strftime('%d.%m.%Y')}"
        else:
            last_payment = f"£{account.last_payment_amount:.2f}"

        # Get category and notes from result
        if result:
            category = result.category
            notes = result.notes
            if result.action_suggested:
                notes += f" | Action: {result.action_suggested}"
        else:
            category = ''
            notes = ''

        # Row data
        row_data = [
            account.client_id,
            account.client_name,
            last_payment,
            account.current if account.current > 0 else '',
            account.days_30 if account.days_30 > 0 else '',
            account.days_60 if account.days_60 > 0 else '',
            account.days_90 if account.days_90 > 0 else '',
            account.total_balance,
            '',  # SMS - empty, user fills in
            '',  # EMAIL - empty, user fills in
            category,
            '',  # ADMIN NOTES - empty, practice manager fills in
            notes  # AI ANALYSIS NOTES
        ]

        # Write row
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border

            # Format numbers
            if col in [4, 5, 6, 7, 8] and value != '':
                cell.number_format = '£#,##0.00'
                cell.alignment = Alignment(horizontal='right')

            # Center and bold the TYPE column (col 11)
            if col == 11:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Apply text wrapping to ADMIN NOTES (col 12) and AI ANALYSIS NOTES (col 13)
            if col in [12, 13]:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        row_num += 1

    # Add conditional formatting for category colors (columns A-K based on TYPE in column K)
    # This makes colors update dynamically when the TYPE value is changed
    last_data_row = row_num - 1  # Last row of data
    data_range = f"A2:K{last_data_row}"

    for category, color in CATEGORY_COLORS.items():
        # Formula checks if column K (TYPE) matches the category
        # Using $K2 to lock the column reference but allow row to vary
        formula = f'$K2="{category}"'
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        rule = FormulaRule(formula=[formula], fill=fill)
        ws.conditional_formatting.add(data_range, rule)

    # Add summary section with Excel formulas (dynamic updates if user edits data)
    row_num += 2
    ws.cell(row=row_num, column=1, value="Summary").font = Font(bold=True, size=14)

    row_num += 1
    ws.cell(row=row_num, column=1, value="Total Accounts:")
    ws.cell(row=row_num, column=2, value=f'=COUNTA(A2:A{last_data_row})')  # Count non-empty IDs
    if len(filtered_accounts) < len(accounts):
        ws.cell(row=row_num, column=3, value=f"({len(accounts) - len(filtered_accounts)} excluded: BAD/INV)")

    row_num += 1
    ws.cell(row=row_num, column=1, value="Total Value:")
    ws.cell(row=row_num, column=2, value=f'=SUM(H2:H{last_data_row})')  # Sum of Total column (H)
    ws.cell(row=row_num, column=2).number_format = '£#,##0.00'

    # Category breakdown with formulas (only for included accounts)
    row_num += 2
    ws.cell(row=row_num, column=1, value="By Category:").font = Font(bold=True)

    # Get list of categories that actually exist in the data
    category_list = set()
    for account in filtered_accounts:
        result = results_map.get(account.client_id)
        if result and result.category:
            category_list.add(result.category)

    for cat in sorted(category_list):
        row_num += 1
        ws.cell(row=row_num, column=1, value=f"  {cat}:")
        # Count formula using COUNTIF on TYPE column (K)
        ws.cell(row=row_num, column=2, value=f'=COUNTIF(K:K,"{cat}")')
        ws.cell(row=row_num, column=3, value="accounts")
        # Sum formula using SUMIF on TYPE column (K) and Total column (H)
        ws.cell(row=row_num, column=4, value=f'=SUMIF(K:K,"{cat}",H:H)')
        ws.cell(row=row_num, column=4).number_format = '£#,##0.00'

        if cat in CATEGORY_COLORS:
            ws.cell(row=row_num, column=1).fill = PatternFill(
                start_color=CATEGORY_COLORS[cat],
                end_color=CATEGORY_COLORS[cat],
                fill_type='solid'
            )

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Add filter (updated to include new column M)
    ws.auto_filter.ref = f"A1:M{len(filtered_accounts) + 1}"

    # Save
    wb.save(output_path)

    return output_path


def create_summary_sheet(
    wb: Workbook,
    accounts: List[DebtAccount],
    results: List[CategorizationResult]
) -> None:
    """Add a summary sheet to the workbook."""
    ws = wb.create_sheet("Summary")

    # Title
    ws.cell(row=1, column=1, value="Debt Analysis Summary").font = Font(bold=True, size=16)
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    # Total stats
    row = 4
    ws.cell(row=row, column=1, value="Total Accounts:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=len(accounts))

    row += 1
    total_value = sum(a.total_balance for a in accounts)
    ws.cell(row=row, column=1, value="Total Value:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=total_value)
    ws.cell(row=row, column=2).number_format = '£#,##0.00'

    # By category
    row += 2
    ws.cell(row=row, column=1, value="By Category").font = Font(bold=True, size=12)

    headers = ['Category', 'Count', 'Value', '% of Total']
    row += 1
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header).font = Font(bold=True)

    # Calculate category breakdown
    results_map = {r.client_id: r for r in results}
    category_data = {}

    for account in accounts:
        result = results_map.get(account.client_id)
        cat = result.category if result else 'Unknown'
        if cat not in category_data:
            category_data[cat] = {'count': 0, 'value': 0}
        category_data[cat]['count'] += 1
        category_data[cat]['value'] += account.total_balance

    for cat, data in sorted(category_data.items()):
        row += 1
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=data['count'])
        ws.cell(row=row, column=3, value=data['value'])
        ws.cell(row=row, column=3).number_format = '£#,##0.00'
        pct = (data['value'] / total_value * 100) if total_value > 0 else 0
        ws.cell(row=row, column=4, value=f"{pct:.1f}%")

    # By aging
    row += 2
    ws.cell(row=row, column=1, value="By Aging").font = Font(bold=True, size=12)

    row += 1
    for col, header in enumerate(['Aging', 'Count', 'Value'], 1):
        ws.cell(row=row, column=col, value=header).font = Font(bold=True)

    aging_data = {
        'Current': {'count': 0, 'value': 0},
        '30 Days': {'count': 0, 'value': 0},
        '60 Days': {'count': 0, 'value': 0},
        '90+ Days': {'count': 0, 'value': 0}
    }

    for account in accounts:
        cat = account.aging_category
        if cat in aging_data:
            aging_data[cat]['count'] += 1
            aging_data[cat]['value'] += account.total_balance

    for aging, data in aging_data.items():
        row += 1
        ws.cell(row=row, column=1, value=aging)
        ws.cell(row=row, column=2, value=data['count'])
        ws.cell(row=row, column=3, value=data['value'])
        ws.cell(row=row, column=3).number_format = '£#,##0.00'

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15


# For testing
if __name__ == '__main__':
    from csv_parser import load_debt_csv
    from debt_analyzer import CategorizationResult

    # Load sample accounts
    test_path = Path(__file__).parent / 'mdebtor.csv'
    accounts = load_debt_csv(test_path)

    # Create mock results (in real use, these come from debt_analyzer)
    mock_results = []
    for i, acc in enumerate(accounts):
        # Assign mock categories for testing
        categories = ['MED', 'MED READY', 'PAY', 'INS', 'SMJ']
        cat = categories[i % len(categories)]

        mock_results.append(CategorizationResult(
            client_id=acc.client_id,
            category=cat,
            confidence='HIGH',
            notes=f'Mock categorization for testing - {cat}',
            action_suggested='Test action'
        ))

    # Generate Excel
    output_path = create_debt_excel(accounts, mock_results)
    print(f"Created test Excel file: {output_path}")
